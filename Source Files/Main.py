import pandas as pd
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import shutil
import os

if os.path.exists(r'result.xlsx'):
  os.remove(r'result.xlsx')
else:
  pass

shutil.copyfile(r'Output_template.xlsx', r'result.xlsx')

df = pd.read_excel (r'Amazon Scrapping Input.xlsx')

options = {}
options['strings_to_formulas'] = False
options['strings_to_urls'] = False

for index, row in df.iterrows():
   
    driver = webdriver.Chrome(executable_path=r"chromedriver_win32\chromedriver.exe")
    driver.get("https://www.amazon.com")
    driver.maximize_window()

    driver.find_element_by_xpath("//input[@id='twotabsearchtextbox']").send_keys(str(row['Producto']).strip())
    driver.find_element_by_xpath('//*[@id="nav-search-submit-button"]').click()
    
    results_initial = driver.find_elements_by_xpath('//*[@class="a-size-mini a-spacing-none a-color-base s-line-clamp-2"]')
    
    product_name_splitted = str(row['Producto']).strip().upper().split()
    results_scrapped_counter = 0
    

    for result in results_initial:
        matches_counter = 0
        item_data = {}
       
        if (int(str(row['Cantidad']).strip()) <= results_scrapped_counter):
            break
        
        for word in product_name_splitted:
            if word.strip() in str(result.find_element_by_xpath('.//*[@class="a-size-medium a-color-base a-text-normal"]').text).upper():
                matches_counter += 1

        if matches_counter == len(product_name_splitted):

            book = load_workbook(r'result.xlsx')
            writer = pd.ExcelWriter(r'result.xlsx', engine='openpyxl', options = options)
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}

            item_data['Nombre'] = str(row['Producto']).strip()
            item_data['Nombre en Amazon'] = [str(result.find_element_by_xpath('.//*[@class="a-size-medium a-color-base a-text-normal"]').text).strip()]

            result.find_element_by_xpath('.//*[@class="a-link-normal a-text-normal"]').send_keys(Keys.CONTROL + Keys.RETURN)
            driver.switch_to.window(driver.window_handles[1])
            
            try:
                item_data['Precio'] = [str(float(str(driver.find_element_by_xpath('//span[@id="price_inside_buybox"]').text).replace(',', '').replace('$', '').strip()) * 3850)]
            except Exception as e:
                item_data['Precio'] = ['Este proveedor no tiene stock']
                print(e)
            item_data['Link'] = [str(driver.current_url)]
            
            
            try:
                specs_table = driver.find_element_by_xpath('//table[@class="a-normal a-spacing-micro"]').find_element_by_xpath('.//tbody').find_elements_by_xpath(".//tr")
                print(len(specs_table))

                for j, spec_row in enumerate(specs_table):
                    row_values = spec_row.find_elements_by_xpath('.//td')
                    if str(row_values[0].text).strip() and str(row_values[1].text).strip():
                        item_data['caracteristica #{}'.format(str(j))] = '{} : {}'.format(str(row_values[0].text).strip(), str(row_values[1].text).strip())
            except:
                print("No specs table")
            
        
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            results_scrapped_counter += 1


            # Append data to results excel
            result_data_frame = pd.DataFrame.from_dict(item_data)
            startrow = writer.sheets['Sheet1'].max_row
            result_data_frame.to_excel(writer, startrow = startrow, sheet_name='Sheet1' , index = False, header = False)
            writer.save()
    
    driver.close()
    driver.quit()